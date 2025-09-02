import os
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import numbers, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from typing import Any, Dict, Generator, List, Tuple
from datetime import datetime
import uuid
from itertools import chain

from cast_exceptions import TablesNotFound, FieldNotFound, RowsNotFound


class ExcelProcessor:
    """
    Класс для обработки Excel-файлов и извлечения данных из таблиц.

    Позволяет:
    - фильтровать данные по определённому полю и значению;
    - удалять лишние поля, оставляя только заданные;
    - извлекать метаданные с листа.
    """

    WANTED_FIELDS: Tuple[str, ...] = ("ФИО", "Должность", "Отдел", "Дата найма", "Зарплата")

    def __init__(self, filepath: str, field_name: str, value: Any, output_path: str = '') -> None:
        """
        :param filepath: Путь к Excel-файлу
        :param field_name: Название поля, по которому будет фильтрация
        :param value: Значение для фильтрации
        :param output_path: Путь для сохранения изменённого файла (по умолчанию — рядом с исходным)
        """
        try:
            self.wb = load_workbook(filepath, data_only=True)
        except PermissionError:
            raise PermissionError("Файл, который вы пытаетесь открыть, занят другим процессом,"
                                  "или у вас нет прав для работы с ним.")

        if output_path and not os.path.splitext(output_path)[1]:
            output_path = f'{output_path}.xlsx'
        self.output_path: str = output_path or os.path.splitext(filepath)[0] + f'_{str(uuid.uuid4())[:8]}.xlsx'
        self.field_name: str = field_name
        self.value: Any = value
        self.fieldnames: List[str] = []

    def process(self):
        """
        Основной метод обработки файла.

        :return: Кортеж из списка метаданных и списка строк таблицы после фильтрации.
        :raises TablesNotFound: Если не удалось найти ни одной таблицы.
        """
        # Создаем новую "книгу" и удаляем лишний лист, который создается по умолчанию
        new_wb = Workbook()
        default_sheet = new_wb.active
        new_wb.remove(default_sheet)

        # Задаем стили для форматирования ячеек
        thin = Side(border_style="thin", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        align = Alignment(horizontal='center', vertical='center')

        # Проходимся по каждому листу
        for sheet_name in self.wb.sheetnames:
            ws = new_wb.create_sheet(title=sheet_name)
            sheet = self.wb[sheet_name]

            # Извлекаем строки самой таблицы и фильтруем их
            raw_data = self._rows_of_table(sheet)
            try:
                first_raw = next(raw_data)
                raw_data = chain([first_raw], raw_data)
            except StopIteration:
                raise TablesNotFound('Ни одной таблицы не найдено')
            tuple_of_fields = next(raw_data)
            self.fieldnames = list([cell.value for cell in tuple_of_fields])
            data = self._format_data(raw_data)
            filtered_data = self._filter_data(data)
            try:
                first_row = next(filtered_data)
                filtered_data = chain([first_row], filtered_data)
            except StopIteration:
                raise RowsNotFound('Ни одной записи с указанным значением не найдено')
            prepared_data = self._del_fields(filtered_data)

            # Извлекаем метаданные
            meta_data = self._get_meta_data(sheet)

            # Сначала записываем метаданные в ячейки по прежним координатам и задаем ячейкам стили
            for row in meta_data:
                for cell in row:
                    if isinstance(cell.value, datetime):
                        ws[cell.coordinate] = cell.value.date()
                        ws[cell.coordinate].number_format = numbers.FORMAT_DATE_DDMMYY
                    ws[cell.coordinate] = cell.value
                    col_letter = get_column_letter(cell.column)
                    ws.column_dimensions[col_letter].width = len(str(cell.value)) + 10

            # Добавляем строку с заголовками таблицы и берем индекс этой строки, чтобы задать соответствующие стили
            fields = list(field for field in self.fieldnames if field in self.WANTED_FIELDS)
            ws.append(fields)
            new_row = ws.max_row
            len_rows = len(fields)
            for cell in ws[new_row]:
                cell.alignment = align
                cell.font = Font(bold=True)

            # Заносим данные таблицы
            for row in prepared_data:
                ws.append(list(cell.value for cell in row.values()))

            # Добавляем стили к таблице. К сожалению, у метода append нет возможности сразу задать стили к строкам.
            for row_idx in range(new_row, ws.max_row + 1):
                for cell in ws[row_idx][:len_rows]:
                    cell.border = border
                    col_letter = get_column_letter(cell.column)
                    ws.column_dimensions[col_letter].width = len(str(cell.value)) + 10
        try:
            # Сохраняем файл
            new_wb.save(self.output_path)
        except PermissionError:
            raise PermissionError("Файл с указанным названием занят другим процессом,"
                                  " или у вас нет прав для работы с ним.")
        return "Готово"

    def _del_fields(self, data) -> Generator[Dict[str, Any], None, None]:
        """
        Удаляет из строк все поля, кроме нужных.
        """
        for row in data:
            for field in self.fieldnames:  # приводим к list, чтобы не менять итерируемое множество
                if field not in self.WANTED_FIELDS:
                    row.pop(field, None)
            yield row

    def _filter_data(self, data: Generator[Dict[str, Any], None, None]) -> Generator[Dict[str, Any], None, None]:
        """
        Фильтрует строки таблицы по значению в указанном поле.
        """
        field = list(filter(lambda x: x.lower() == self.field_name.lower(), self.fieldnames))
        if field:
            for row in data:
                if str(row[field[0]].value) == self.value:
                    yield row
        else:
            raise FieldNotFound('Указанного поля в таблице нет')

    def _rows_of_table(self, sheet: Worksheet) -> Generator[Any, None, None]:
        """
        Извлекает строки таблицы с листа, игнорируя неполные строки.
        """
        is_table = False
        for row in sheet.iter_rows():
            is_table = is_table or any(cell.value in self.WANTED_FIELDS for cell in row)
            if is_table:
                yield row

    def _format_data(self, data) -> Generator[Dict[str, Any], None, None]:
        """
        Преобразует список строк таблицы в список словарей.
        """
        for row in data:
            yield {field_name: cell for field_name, cell in zip(self.fieldnames, row)}

    def _get_meta_data(self, sheet: Worksheet) -> Generator[Any, None, None]:
        """
        Извлекает метаданные с листа.

        Под метаданными понимаются строки, которые не являются частью таблицы,
        а содержат служебную информацию.
        """
        is_table = False
        for row in sheet.iter_rows():
            is_table = is_table or any(cell.value in self.WANTED_FIELDS for cell in row)
            if not is_table:
                yield row
